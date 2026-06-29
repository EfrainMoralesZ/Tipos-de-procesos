def inicializar_datos_externos():
    """Copia los archivos de datos desde el paquete interno a la carpeta externa si no existen."""
    import shutil
    if getattr(sys, 'frozen', False):
        # Carpeta junto al .exe
        datos_externos = os.path.join(os.path.dirname(sys.executable), 'datos')
        os.makedirs(datos_externos, exist_ok=True)
        # Carpeta interna del paquete
        datos_internos = os.path.join(sys._MEIPASS, 'datos')
        archivos = [
            'codigos_cumple.xlsx',
            'base_general.json',
            'codigos_cumple.json',
            'config.json'
        ]
        for archivo in archivos:
            externo = os.path.join(datos_externos, archivo)
            interno = os.path.join(datos_internos, archivo)
            if not os.path.exists(externo) and os.path.exists(interno):
                try:
                    shutil.copy2(interno, externo)
                    print(f"Archivo copiado a datos externos: {archivo}")
                except Exception as e:
                    print(f"Error copiando {archivo}: {e}")

# --- Funciones utilitarias centralizadas ---
def cargar_excel(path, columnas_requeridas=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"No se encontró el archivo: {path}")
    engine = 'xlrd' if path.lower().endswith('.xls') else 'openpyxl'
    df = pd.read_excel(path, engine=engine)
    if columnas_requeridas and not set(columnas_requeridas).issubset(df.columns):
        raise ValueError(f"Faltan columnas requeridas en {path}")
    return df

def cargar_json(path, columnas_requeridas=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"No se encontró el archivo: {path}")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    df = pd.DataFrame(data)
    if columnas_requeridas and not set(columnas_requeridas).issubset(df.columns):
        raise ValueError(f"Faltan columnas requeridas en {path}")
    return df

def limpiar_items(df, columna):
    items = pd.to_numeric(df[columna], errors='coerce')
    items = items.dropna().astype(int)
    return list(set(items))

def modificar_tipo_proceso(row, normas_adherible, normas_costura):
    norma_val = str(row['NORMA'])
    tipo = str(row['TIPO DE PROCESO'])
    if 'NOM004TEXX' in tipo or 'TEXX' in norma_val:
        return 'ADHERIBLE'
    if 'NOM004' in tipo or '004' in tipo or 'NOM-004-SE-2021' in norma_val:
        return 'COSTURA'
    if 'NOM020INS' in norma_val:
        return 'ADHERIBLE'
    if any(n in norma_val for n in normas_adherible):
        return 'ADHERIBLE'
    if any(n in norma_val for n in normas_costura):
        return 'COSTURA'
    if norma_val == '0':
        return 'SIN NORMA'
    if norma_val == 'N/D':
        return ''
    return tipo

def modificar_norma(norma_val):
    if str(norma_val) == '0':
        return 'SIN NORMA'
    elif str(norma_val) == 'N/D':
        return ''
    return norma_val

def modificar_criterio(crit_val):
    crit = str(crit_val).strip().upper()
    if 'NO CUMPLE' in crit:
        return crit_val
    if any(palabra in crit for palabra in ['CUMPLE', 'C']):
        return 'CUMPLE'
    return crit_val

import os
import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import sys
import json
from Editor_Codigos import EditorCodigos
from Formato import exportar_excel
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as pdf_canvas
from io import BytesIO
from reportlab.lib.utils import ImageReader
import Rutas  # Debe estar en la misma carpeta que ProcesosV2.py
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font


# Configuración de rutas para .py y .exe
if getattr(sys, 'frozen', False):
    # Cuando está compilado en .exe
    BASE_PATH = sys._MEIPASS
else:
    # Cuando se ejecuta desde Python
    BASE_PATH = os.path.dirname(os.path.abspath(__file__))

# Archivos de configuración
# Archivos de configuración centralizados
CONFIG_FILE = Rutas.archivo_datos("config.json")
ARCHIVOS_PROCESADOS_FILE = Rutas.archivo_datos("archivos_procesados.json")
CODIGOS_CUMPLE_FILE = Rutas.archivo_datos("codigos_cumple.xlsx")
CODIGOS_JSON_FILE = Rutas.archivo_datos("codigos_cumple.json")
BASE_GENERAL_JSON = Rutas.archivo_datos("base_general.json")


def asegurar_excel_vacio(ruta, columnas):
    """Crea un archivo Excel vacío con las columnas especificadas si no existe"""
    if not os.path.exists(ruta):
        import pandas as pd
        df = pd.DataFrame(columns=columnas)
        df.to_excel(ruta, index=False)
        print(f"✅ Archivo Excel creado vacío: {ruta}")

# Configuración de Rutas integrada
def configurar_rutas(parent=None):
    try:
        import Configurar  # El módulo debe estar en la misma carpeta que ProcesosV2.py
        # Pasar el parent principal si existe
        if parent is not None:
            Configurar.configurar_rutas(parent)
        else:
            Configurar.configurar_rutas()
    except Exception as e:
        messagebox.showerror("❌ Error", f"No se pudo abrir la configuración:\n{e}")

def obtener_archivos_procesados():
    asegurar_json(ARCHIVOS_PROCESADOS_FILE, [])
    try:
        with open(ARCHIVOS_PROCESADOS_FILE, 'r', encoding='utf-8') as f:
            datos = json.load(f)
            return datos if isinstance(datos, list) else []
    except Exception as e:
        print(f"❌ Error cargando archivos procesados: {e}")
        return []

def cargar_archivos_procesados():
    """Carga la lista de archivos procesados, crea el JSON si no existe"""
    asegurar_json(ARCHIVOS_PROCESADOS_FILE, [])
    try:
        with open(ARCHIVOS_PROCESADOS_FILE, 'r', encoding='utf-8') as f:
            datos = json.load(f)
            return datos if isinstance(datos, list) else []
    except Exception as e:
        print(f"❌ Error cargando archivos procesados: {e}")
        return []

def registrar_archivo_procesado(nombre_archivo, fecha_proceso):
    """Registra un archivo procesado en el sistema de estadísticas"""
    try:
        archivos_procesados = obtener_archivos_procesados()
        # Evitar duplicados
        if any(a["nombre"] == nombre_archivo for a in archivos_procesados):
            print(f"ℹ️ Archivo ya registrado: {nombre_archivo}")
            return
        archivo_info = {
            "nombre": nombre_archivo,
            "fecha_proceso": fecha_proceso,
            "fecha_archivo": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        archivos_procesados.append(archivo_info)
        with open(ARCHIVOS_PROCESADOS_FILE, 'w', encoding='utf-8') as f:
            json.dump(archivos_procesados, f, indent=4, ensure_ascii=False)
        print(f"✅ Archivo registrado correctamente: {nombre_archivo}")
    except Exception as e:
        print(f"❌ Error registrando archivo: {e}")

# OBTENER ESTADISTICAS DE ARCHIVOS
def obtener_estadisticas_archivos():
    """Obtiene estadísticas de archivos procesados"""
    try:
        if os.path.exists(ARCHIVOS_PROCESADOS_FILE):
            with open(ARCHIVOS_PROCESADOS_FILE, 'r', encoding='utf-8') as f:
                archivos = json.load(f)
            return {
                "total_archivos": len(archivos),
                "archivos_recientes": archivos[-5:] if len(archivos) > 5 else archivos,
                "ultimo_proceso": archivos[-1]["fecha_proceso"] if archivos else "Ninguno"
            }
        else:
            return {
                "total_archivos": 0,
                "archivos_recientes": [],
                "ultimo_proceso": "Ninguno"
            }
    except Exception as e:
        print(f"[ERROR] Error obteniendo estadísticas: {e}")
        return {
            "total_archivos": 0,
            "archivos_recientes": [],
            "ultimo_proceso": "Error"
        }

def asegurar_json(ruta, contenido_inicial):
    """Crea un archivo JSON solo si no existe o está completamente vacío (0 bytes). Nunca sobrescribe si ya tiene datos."""
    if not os.path.exists(ruta) or os.path.getsize(ruta) == 0:
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(contenido_inicial, f, indent=4, ensure_ascii=False)

# CARGAR CONFIGURACION DE RUTAS
def cargar_configuracion():
    """Carga la configuración desde config.json"""
    # Crear archivos si no existen
    asegurar_json(CONFIG_FILE, {"rutas": {"base_general": "", "codigos_cumple": ""}})
    asegurar_json(ARCHIVOS_PROCESADOS_FILE, [])
    asegurar_json(CODIGOS_JSON_FILE, [])
    asegurar_json(BASE_GENERAL_JSON, [])

    # Crear XLSX vacío si no existe
    asegurar_excel_vacio(CODIGOS_CUMPLE_FILE, columnas=["ITEM", "CRITERIO", "OBSERVACIONES"])

    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error al cargar configuración: {e}")
        return {"rutas": {"base_general": "", "codigos_cumple": ""}}

def guardar_configuracion(config):
    """Guarda la configuración en config.json"""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Error al guardar configuración: {e}")

# FUNCION PARA ACTUALIZAR CODIGOS 
def abrir_editor_codigos(parent):
    """Abre el editor de códigos según la configuración de rutas"""
    try:
        config = cargar_configuracion()
        if not config:
            messagebox.showerror("Error", "No se pudo cargar la configuración")
            return None

        # Obtener rutas de archivos
        rutas = config.get("rutas", {})
        archivo_codigos = rutas.get("codigos_cumple", "")
        archivo_json = ""
        
        if archivo_codigos:
            archivo_json = Rutas.archivo_datos("codigos_cumple.json")

        # Validar que existan los archivos
        if os.path.exists(archivo_codigos) and os.path.exists(archivo_json):
            editor = EditorCodigos(parent, archivo_codigos, archivo_json)
            return editor
        else:
            messagebox.showwarning(
                "Advertencia",
                "Primero debe configurar los archivos en Configuración de Rutas.\n"
                f"Archivos esperados:\n{archivo_codigos}\n{archivo_json}"
            )
            return None

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al abrir el editor de códigos:\n{e}")
        return None

# FUNCIONES AUXILIARES
def modificar_tipo_proceso(row, normas_adherible, normas_costura):
    norma = str(row['NORMA']).upper().strip()
    descripcion = str(row['DESCRIPCION']).upper().strip()
    tipo = str(row.get('TIPO DE PROCESO', '')).upper().strip()

    # Excepción NOM-004: calcetines, medias, ropa interior
    excepciones_nom004 = ['CALCETIN', 'CALCETINES', 'MEDIA', 'MEDIAS', 'SOCKS', 'ROPA INTERIOR', 
                          'PANTIMEDIAS', 'BANDA PARA LA CABEZA','MUÑEQUERAS', 'CALCETAS']
    if norma in ['NOM-004-SE-2021', 'NOM004']:
        for palabra in excepciones_nom004:
            if palabra in descripcion:
                return 'ADHERIBLE'

    # Normas adheribles
    if norma in normas_adherible:
        return 'ADHERIBLE'

    # Normas de costura
    if norma in normas_costura:
        return 'COSTURA'

    # Si ya tiene tipo asignado
    if tipo:
        return tipo

    # Default
    return 'SIN NORMA'

#  FUNCION PARA GENERAR EL TIPO DE PROCESO

def procesar_reporte(reporte_path):
    # REGISTRAR ARCHIVO PROCESADO
    nombre_archivo = os.path.basename(reporte_path)
    fecha_proceso = datetime.now().strftime("%Y-%m-%d")
    registrar_archivo_procesado(nombre_archivo, fecha_proceso)

    # SE CREA LA BARRA DE PROGRESO EN EL FRAME PRINCIPAL (LADO DERECHO)
    try:
        # Crear barra de progreso
        barra = BarraProgreso(frame, texto="Procesando...", ancho=220, posicion="derecha")

        try:
            # Limpiar widgets existentes si existen
            for widget_name in ['progress_label', 'progress_bar', 'percent_label']:
                if widget_name in globals():
                    widget = globals()[widget_name]
                    if widget is not None and hasattr(widget, 'destroy'):
                        try:
                            widget.destroy()
                        except:
                            pass
        except Exception:
            pass

        # Cargar archivos usando funciones centralizadas
        df_base = cargar_json(Rutas.archivo_datos("base_general.json"), columnas_requeridas=["EAN", "CODIGO FORMATO"])
        df_codigos_cumple = cargar_json(Rutas.archivo_datos("codigos_cumple.json"), columnas_requeridas=["ITEM", "OBSERVACIONES", "CRITERIO"])
        df_reporte = cargar_excel(reporte_path)

        # Guardar un JSON de depuración del reporte de mercancía en la carpeta 'datos'
        try:
            datos_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'datos')
            os.makedirs(datos_dir, exist_ok=True)
            debug_json_path = os.path.join(
                datos_dir,
                f'debug_reporte_mercancia_{os.path.splitext(os.path.basename(reporte_path))[0]}.json'
            )
            df_reporte.to_json(debug_json_path, orient='records', force_ascii=False, indent=4)
            print(f"[INFO] JSON de depuración generado: {debug_json_path}")
        except Exception as e:
            print(f"[ERROR] No se pudo guardar el JSON de depuración: {e}")

        # Detectar columnas clave
        if 'Número de Parte' in df_reporte.columns:
            num_parte_col = 'Número de Parte'
            desc_col = 'Desc. Pedimento'
            norma_col = 'Normas'
            criterio_col = 'CRITERIO'
        elif any(col.strip().lower() in ['num. parte', 'num.parte', 'numero de parte'] for col in df_reporte.columns):
            for col in df_reporte.columns:
                if col.strip().lower() in ['num. parte', 'num.parte', 'numero de parte']:
                    num_parte_col = col
                    break
            for col in df_reporte.columns:
                if col.strip().lower() == 'descripción agente aduanal':
                    desc_col = col
                    break
            norma_col = 'NOMs'
            criterio_col = 'CRITERIO'
        else:
            raise ValueError("No se encontró ninguna columna de NUM. PARTE válida en el reporte")

        # Limpiar y obtener lista de items únicos
        items = limpiar_items(df_reporte, num_parte_col)
        total = len(items)

        # Filtrar DataFrames por los items
        # Para merge, todos los campos clave deben ser string
        df_items = pd.DataFrame({num_parte_col: items})
        df_items[num_parte_col] = df_items[num_parte_col].astype(str)
        df_base['EAN'] = df_base['EAN'].astype(str)
        df_codigos_cumple['ITEM'] = df_codigos_cumple['ITEM'].astype(str)
        df_reporte[num_parte_col] = df_reporte[num_parte_col].astype(str).str.strip()

        # Merge para obtener toda la información en un solo DataFrame
        df_result = df_items.merge(df_base, left_on=num_parte_col, right_on='EAN', how='left')
        df_result = df_result.merge(df_reporte[[num_parte_col, desc_col, norma_col]], on=num_parte_col, how='left')
        df_result = df_result.merge(df_codigos_cumple[["ITEM", "OBSERVACIONES", "CRITERIO"]], left_on=num_parte_col, right_on="ITEM", how="left")

        # Renombrar y construir columnas finales
        df_result['ITEM'] = pd.to_numeric(df_result[num_parte_col], errors="coerce").astype('Int64')
        df_result['TIPO DE PROCESO'] = df_result['CODIGO FORMATO'].fillna('')
        df_result['NORMA'] = df_result[norma_col].fillna('')
        df_result['DESCRIPCION'] = df_result[desc_col].fillna('')
        # Criterio: si OBSERVACIONES contiene 'CUMPLE', poner 'CUMPLE', si no, usar CRITERIO
        df_result['CRITERIO'] = df_result.apply(lambda row: 'CUMPLE' if str(row['OBSERVACIONES']).upper().strip() == 'CUMPLE' else (row['CRITERIO'] if pd.notna(row['CRITERIO']) else ''), axis=1)

        # Reglas de negocio
        normas_adherible = [
            'NOM-050-SCFI-2004', 'NOM-121-SCFI-2004',
            'NOM-015-SCFI-2007', 'NOM-050-SCFI-2004',
            'NOM-024-SCFI-2013', 'NOM-141-SSA1/SCFI-2012',
            'NOM004TEXX', 'NOM020INS', 'NOM-115-STPS-2009','NOM-189-SSA1/SCFI-2018'
        ]
        normas_costura = ['NOM-004-SE-2021', 'NOM-020-SCFI-1997', 'NOM004', 'NOM020']
        normas_validas = ['003','NOM-004-SE-2021','008','NOM-015-SCFI-2007','020','NOM-020-SCFI-1997',
                          'NOM-024-SCFI-2013','035','NOM-050-SCFI-2004','051','116','NOM-141-SSA1/SCFI-2012','142','173','185','186','NOM-189-SSA1/SCFI-2018','192','199','235','NOM-115-STPS-2009','NOM-121-SCFI-2004']

        # DEBUG: columna de norma detectada y valor crudo en el reporte
        print(f"\n[DEBUG] Columna de norma detectada: '{norma_col}'")
        item_debug = '4479813'
        raw = df_reporte[df_reporte[num_parte_col].astype(str).str.strip() == item_debug]
        print(f"[DEBUG] Filas en df_reporte para item {item_debug}:")
        print(raw[[num_parte_col, norma_col]].to_string() if norma_col in df_reporte.columns else f"  columna '{norma_col}' NO existe en df_reporte")
        merged_row = df_result[df_result[num_parte_col].astype(str).str.strip() == item_debug]
        print(f"[DEBUG] Fila en df_result tras merge para item {item_debug}:")
        print(merged_row[[num_parte_col, norma_col, 'NORMA']].to_string() if norma_col in df_result.columns else merged_row[[num_parte_col, 'NORMA']].to_string())

        # DEBUG: estado antes de aplicar reglas
        print("\n===== DEBUG: ANTES DE APLICAR REGLAS =====")
        print(df_result[['ITEM', 'NORMA', 'TIPO DE PROCESO', 'CRITERIO']].to_string())
        print("==========================================\n")

        # Aplicar reglas
        df_result['TIPO DE PROCESO'] = df_result.apply(lambda row: modificar_tipo_proceso(row, normas_adherible, normas_costura), axis=1)
        df_result['NORMA'] = df_result['NORMA'].apply(modificar_norma)
        df_result['CRITERIO'] = df_result['CRITERIO'].apply(modificar_criterio)

        # DEBUG: estado después de aplicar reglas
        print("\n===== DEBUG: DESPUÉS DE APLICAR REGLAS =====")
        print(df_result[['ITEM', 'NORMA', 'TIPO DE PROCESO', 'CRITERIO']].to_string())
        print("============================================\n")

        # Vectorizar reglas adicionales sobre el DataFrame
        df_result['TIPO DE PROCESO'] = df_result['TIPO DE PROCESO'].astype(str).str.strip()
        df_result['NORMA'] = df_result['NORMA'].astype(str).str.strip()
        df_result['CRITERIO'] = df_result['CRITERIO'].astype(str).str.strip().str.upper()

        df_result.loc[~df_result['NORMA'].isin(normas_validas), 'TIPO DE PROCESO'] = 'SIN NORMA'
        df_result.loc[df_result['NORMA'].isin(['', '0']), 'NORMA'] = 'SIN NORMA'
        df_result.loc[(df_result['TIPO DE PROCESO'] == '') | \
                     ((df_result['TIPO DE PROCESO'] == '0') & (df_result['NORMA'] == '0')) | \
                     ((df_result['TIPO DE PROCESO'] == '') & (df_result['NORMA'] == '')), ['TIPO DE PROCESO', 'NORMA']] = ['SIN NORMA', 'SIN NORMA']
        df_result.loc[df_result['CRITERIO'].str.contains('CUMPLE', na=False), ['TIPO DE PROCESO', 'CRITERIO']] = ['CUMPLE', '']

        # DEBUG: estado final después de reglas vectorizadas
        print("\n===== DEBUG: RESULTADO FINAL =====")
        print(df_result[['ITEM', 'NORMA', 'TIPO DE PROCESO', 'CRITERIO']].to_string())
        print("==================================\n")

        #Cambia la forma en la que se imprime el resultado en la columna "CRITERIO" poniendo todo en REVISADO
        # Si CRITERIO dice "REVISADO", reemplázalo por el texto de OBSERVACIONES
        df_result.loc[df_result['CRITERIO'].str.upper() == 'REVISADO', 'CRITERIO'] = df_result['OBSERVACIONES']


        # df_result.loc[(df_result['NORMA'].isin(['NOM-050-SCFI-2004', 'NOM-015-SCFI-2007'])) & \
        #              ~df_result['CRITERIO'].str.contains('CUMPLE', na=False), 'TIPO DE PROCESO'] = 'ADHERIBLE'

        # --- Asignar columna PIEZAS sumando la 'Cantidad Factura' por cada ITEM (número de parte) ---
        # Buscar la columna correcta de cantidad facturada
        cantidad_cols = ['Cantidad Factura', 'Cantidad Facturada', 'I', 'AQ','Cant. comercial']
        cant_col = None
        for col in cantidad_cols:
            if col in df_reporte.columns:
                cant_col = col
                break

        if cant_col:
            # Normalizar ambas columnas como texto sin espacios
            df_reporte[num_parte_col] = df_reporte[num_parte_col].astype(str).str.strip()
            df_result['ITEM'] = df_result['ITEM'].astype(str).str.strip()

            # Agrupar el reporte para sumar piezas por número de parte
            df_piezas = (
                df_reporte.groupby(num_parte_col, as_index=False)[cant_col]
                .sum()
                .rename(columns={cant_col: 'PIEZAS'})
            )

            # Merge entre df_result y df_piezas
            df_result = df_result.merge(
                df_piezas,
                left_on='ITEM',
                right_on=num_parte_col,
                how='left'
            )

            # Limpiar resultados
            df_result['PIEZAS'] = df_result['PIEZAS'].fillna(0).astype(int)

            # Quitar la columna auxiliar si quedó duplicada
            if num_parte_col in df_result.columns:
                df_result.drop(columns=[num_parte_col], inplace=True)
        else:
            df_result['PIEZAS'] = 0


        # PREVIO: buscar la observación en codigos_cumple.json según ITEM
        df_result['PREVIO'] = df_result.apply(
            lambda row: df_codigos_cumple.loc[df_codigos_cumple['ITEM'] == str(row['ITEM']), 'OBSERVACIONES'].values[0]
            if not df_codigos_cumple.loc[df_codigos_cumple['ITEM'] == str(row['ITEM']), 'OBSERVACIONES'].empty else 0,
            axis=1
        )

        # --- Columnas que se imprimen en el tipo de proceso --- #
        df_result = df_result[['ITEM', 'TIPO DE PROCESO', 'NORMA', 'CRITERIO', 'DESCRIPCION', 'PIEZAS', 'PREVIO']]
        # Asegurar que ITEM se exporte como entero en el Excel
        df_result['ITEM'] = df_result['ITEM'].astype('Int64')

        # ---  Eliminar duplicados asegurando que ITEM sea único ---
        df_result = df_result.drop_duplicates(subset=["ITEM"], keep="first").reset_index(drop=True)

        barra.finalizar("¡Completado!")

        # # --- Crear hoja de Códigos Actualizados ---
        # df_codigos_actualizados = df_codigos_cumple[['ITEM', 'OBSERVACIONES', 'CRITERIO']].copy()
        # df_codigos_actualizados['ITEM'] = pd.to_numeric(df_codigos_actualizados['ITEM'], errors='coerce').astype('Int64')

        # # Filtrar ITEM que están en df_result
        # df_codigos_actualizados = df_codigos_actualizados[
        #     df_codigos_actualizados['ITEM'].isin(df_result['ITEM'])
        # ].drop_duplicates(subset=['ITEM'], keep='first').reset_index(drop=True)

        # # 🪄 Si la columna OBSERVACIONES está vacía, copiar valor de CRITERIO
        # df_codigos_actualizados['OBSERVACIONES'] = df_codigos_actualizados.apply(
        #     lambda row: row['CRITERIO'] if pd.isna(row['OBSERVACIONES']) or row['OBSERVACIONES'] == '' else row['OBSERVACIONES'],
        #     axis=1
        # )


        # # Si algún ITEM no está en codigos_cumple, agregarlo vacío para mantener correspondencia 1 a 1
        # items_faltantes = df_result[~df_result['ITEM'].isin(df_codigos_actualizados['ITEM'])]['ITEM']
        # if not items_faltantes.empty:
        #     df_faltantes = pd.DataFrame({
        #         'ITEM': items_faltantes,
        #         'OBSERVACIONES': [''] * len(items_faltantes),
        #         'CRITERIO': [''] * len(items_faltantes)
        #     })
        #     df_codigos_actualizados = pd.concat([df_codigos_actualizados, df_faltantes], ignore_index=True)

        # # Ordenar igual que df_result
        # df_codigos_actualizados = df_codigos_actualizados.sort_values(by='ITEM').reset_index(drop=True)

        # Guardar archivo final
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            title="Guardar archivo TIPO DE PROCESO",
            initialfile="TIPO DE PROCESO.xlsx"
        )

        if save_path:
            try:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    # Hoja 1: Tipo de proceso
                    df_result.to_excel(writer, index=False, sheet_name="TIPO DE PROCESO")
                    # Hoja 2: Codigos Actualizados
                    # df_codigos_actualizados.to_excel(writer, index=False, sheet_name="Codigos Actualizados")

                    # --- Formato de encabezados y columnas ---
                    workbook = writer.book
                    header_fill = PatternFill(start_color='4FADEA', end_color='4FADEA', fill_type='solid')
                    header_font = Font(color='000000')
                    header_align = Alignment(horizontal='center', vertical='center')

                    def aplicar_estilo(ws, df):
                        # Encabezado azul centrado
                        for cell in list(ws[1]):
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_align

                        # Ajuste de ancho dinámico
                        for i, col in enumerate(df.columns, 1):
                            col_letter = get_column_letter(i)
                            try:
                                max_length = max(
                                    df[col].astype(str).map(len).max(),
                                    len(str(col))
                                )
                            except Exception:
                                max_length = len(str(col))
                            adjusted_width = max(10, min(60, int(max_length * 1.2) + 2))
                            ws.column_dimensions[col_letter].width = adjusted_width

                        # Congelar encabezado
                        ws.freeze_panes = 'A2'

                    # Aplicar estilo a ambas hojas
                    ws1 = writer.sheets.get("TIPO DE PROCESO")
                    ws2 = writer.sheets.get("Codigos Actualizados")
                    if ws1 is not None:
                        aplicar_estilo(ws1, df_result)
                    # if ws2 is not None:
                    #     aplicar_estilo(ws2, df_codigos_actualizados)

                messagebox.showinfo("Éxito", f"Archivo guardado correctamente:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{e}")
        else:
            messagebox.showwarning("Cancelado", "No se guardó el archivo.")

    except Exception as e: messagebox.showerror("Error", f"Ocurrió un problema:\n{e}")

def seleccionar_reporte():
    
    ruta = filedialog.askopenfilename(
        title="Seleccionar REPORTE DE MERCANCIA",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")]
    )
    if ruta:
        procesar_reporte(ruta)

#  VENTANA DEL DASHBOARD
def mostrar_estadisticas():
    """Llama al archivo Dashboard.py para mostrar el dashboard externo"""
    try:
        import Dashboard
        Dashboard.main()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el dashboard:\n{e}")

#  FUNCION PARA LA BARRA DE PROGRESO 
class BarraProgreso:
    def __init__(self, frame, texto="Procesando...", ancho=250, posicion="derecha"):
        """
        frame: contenedor donde se mostrará la barra
        texto: texto de la barra
        ancho: longitud de la barra
        posicion: "derecha" o "izquierda"
        """
        self.frame = frame
        self.ancho = ancho
        self.var = tk.DoubleVar()
        
        self.lbl = tk.Label(frame, text=texto, font=("INTER", 10, "bold"), bg="#FFFFFF", fg="#282828")
        self.percent_lbl = tk.Label(frame, text="0%", font=("INTER", 10, "bold"), bg="#FFFFFF", fg="#282828")
        self.bar = ttk.Progressbar(frame, variable=self.var, maximum=100, length=self.ancho)
        
        # Guardar posición
        self.posicion = posicion
        self._colocar_widgets()
        frame.update()

    def _colocar_widgets(self):
        """Coloca los widgets según la posición deseada."""
        if self.posicion == "derecha":
            anchor = "se"
            x_offset = -20
        else:  # izquierda
            anchor = "sw"
            x_offset = 20

        # Barra y etiquetas
        self.bar.place(relx=1.0 if self.posicion=="derecha" else 0.0, rely=1.0, x=x_offset, y=-40, anchor=anchor)
        self.lbl.place(relx=1.0 if self.posicion=="derecha" else 0.0, rely=1.0, x=x_offset, y=-60, anchor=anchor)
        self.percent_lbl.place(relx=1.0 if self.posicion=="derecha" else 0.0, rely=1.0, x=x_offset, y=-20, anchor=anchor)

    def actualizar(self, valor, texto=None):
        try:
            self.var.set(valor)
            if texto and hasattr(self, 'lbl') and self.lbl.winfo_exists():
                self.lbl.config(text=texto)
            if hasattr(self, 'percent_lbl') and self.percent_lbl.winfo_exists():
                self.percent_lbl.config(text=f"{int(valor)}%")
            if hasattr(self, 'frame') and self.frame.winfo_exists():
                self.frame.update()
        except Exception as e:
            print(f"Error actualizando barra de progreso: {e}")

    def finalizar(self, mensaje="¡Completado!"):
        try:
            self.var.set(100)
            if hasattr(self, 'lbl') and self.lbl.winfo_exists():
                self.lbl.config(text=mensaje)
            if hasattr(self, 'percent_lbl') and self.percent_lbl.winfo_exists():
                self.percent_lbl.config(text="100%")
            if hasattr(self, 'frame') and self.frame.winfo_exists():
                self.frame.update()
                # Ocultar widgets después de un tiempo
                self.frame.after(800, self._ocultar)
        except Exception as e:
            print(f"Error finalizando barra de progreso: {e}")

    def _ocultar(self):
        try:
            if hasattr(self, 'bar') and self.bar.winfo_exists():
                self.bar.place_forget()
            if hasattr(self, 'lbl') and self.lbl.winfo_exists():
                self.lbl.place_forget()
            if hasattr(self, 'percent_lbl') and self.percent_lbl.winfo_exists():
                self.percent_lbl.place_forget()
        except Exception as e:
            print(f"Error ocultando widgets: {e}")

#  VENTANA PRINCIPAL 
root = tk.Tk()
root.title("GENERADOR DE TIPO DE PROCESO")
root.geometry("700x450")
root.configure(bg="#FFFFFF")


# --- Estilo global ---
if __name__ == "__main__":
    # Inicializar datos externos si es .exe
    inicializar_datos_externos()
    
    # Configurar estilo global
    archivos_procesados = cargar_archivos_procesados()
    style = ttk.Style()
    style.theme_use('clam')
    
    # Frame principal con fondo blanco
    frame = tk.Frame(root, bg="#FFFFFF")
    frame.pack(expand=True, fill="both", padx=30, pady=30)

    # --- Header con logo y título mejorado ---
    header_frame = tk.Frame(frame, bg="#FFFFFF")
    header_frame.pack(fill="x", pady=(0, 25))

    # Contenedor para logo y título
    header_content = tk.Frame(header_frame, bg="#FFFFFF")
    header_content.pack(fill="x", pady=(0, 15))

    # Logo en la parte izquierda
    logo_frame = tk.Frame(header_content, bg="#FFFFFF")
    logo_frame.pack(side="left", padx=(0, 20))

    try:
        logo_path = os.path.join(BASE_PATH, "img", "logo.png")
        if os.path.exists(logo_path):
            logo_img_raw = Image.open(logo_path).resize((80, 50), Image.Resampling.LANCZOS)
            logo_img = ImageTk.PhotoImage(logo_img_raw)
            logo_label = tk.Label(logo_frame, image=logo_img, bg="#FFFFFF")
            logo_label.image = logo_img
            logo_label.pack()
    except Exception as e:
        print(f"Error cargando el logo: {e}")

    # Título a la derecha del logo
    title_frame = tk.Frame(header_content, bg="#FFFFFF")
    title_frame.pack(side="left", fill="both", expand=True)

    label_titulo = tk.Label(
        title_frame, 
        text="INSPECCIÓN DE CUMPLIMIENTO\nNORMATIVO AL ARRIBO",
        font=("Inter", 18, "bold"),
        fg="#282828", 
        bg="#FFFFFF", 
        justify="left"
    )
    label_titulo.pack(anchor="w", pady=(0, 8))

    label_sub = tk.Label(
        title_frame, 
        text="Sistema integral para la gestión de procesos normativos",
        font=("Inter", 10),
        fg="#4B4B4B", 
        bg="#FFFFFF",
        justify="left"
    )
    label_sub.pack(anchor="w")

    # Separador decorativo
    separator = tk.Frame(header_frame, height=3, bg="#ECD925")
    separator.pack(fill="x")

    # --- Contenido principal: Botones en disposición moderna ---
    content_frame = tk.Frame(frame, bg="#FFFFFF")
    content_frame.pack(fill="both", expand=True, pady=(30, 20))

    # Configurar estilos de botones mejorados
    style.configure('Primary.TButton', 
                   background='#ECD925', 
                   foreground='#282828', 
                   font=('Inter', 11, 'bold'),
                   borderwidth=0,
                   padding=(15, 12),
                   focuscolor='none')
    style.map('Primary.TButton',
             background=[('active', '#D6BC00')],
             foreground=[('active', '#282828')])
    
    style.configure('Secondary.TButton', 
                   background='#282828', 
                   foreground='#FFFFFF', 
                   font=('Inter', 11, 'bold'),
                   borderwidth=0,
                   padding=(15, 12),
                   focuscolor='none')
    style.map('Secondary.TButton',
             background=[('active', '#1A1A1A')],
             foreground=[('active', '#FFFFFF')])

    # Contenedor principal de botones
    main_buttons_container = tk.Frame(content_frame, bg="#FFFFFF")
    main_buttons_container.pack(expand=True)

    # Primera fila: 3 botones principales
    top_row = tk.Frame(main_buttons_container, bg="#FFFFFF")
    top_row.pack(pady=(0, 15))

    top_buttons = [
        ("⚙️ CONFIGURAR", lambda: configurar_rutas(main_buttons_container)),
        ("📊 REPORTE", seleccionar_reporte),
        ("📋 EDITOR", lambda: abrir_editor_codigos(main_buttons_container))
    ]

    for texto, comando in top_buttons:
        btn_container = tk.Frame(top_row, bg="#FFFFFF", padx=8)
        btn_container.pack(side="left", expand=True)
        
        btn = ttk.Button(btn_container, text=texto, command=comando, 
                        style='Primary.TButton', width=20)
        btn.pack(fill="x", ipady=6)

    # Segunda fila: 2 botones
    bottom_row = tk.Frame(main_buttons_container, bg="#FFFFFF")
    bottom_row.pack(pady=(15, 0))

    # Centrar los dos botones de la segunda fila
    center_frame = tk.Frame(bottom_row, bg="#FFFFFF")
    center_frame.pack()

    bottom_buttons = [
        ("📈 DASHBOARD", mostrar_estadisticas, 'Primary.TButton'),
        ("🚪 CERRAR", lambda: root.destroy() if messagebox.askokcancel(
            "Salir", "¿Está seguro que desea cerrar la aplicación?") else None, 'Secondary.TButton')
    ]

    for texto, comando, estilo in bottom_buttons:
        btn_container = tk.Frame(center_frame, bg="#FFFFFF", padx=12)
        btn_container.pack(side="left", expand=True)
        
        btn = ttk.Button(btn_container, text=texto, command=comando, 
                        style=estilo, width=20)
        btn.pack(fill="x", ipady=6)

    # --- Footer mejorado ---
    footer_frame = tk.Frame(frame, bg="#FFFFFF")
    footer_frame.pack(fill="x", pady=(25, 0))

    # Separador del footer
    footer_separator = tk.Frame(footer_frame, height=1, bg="#E0E0E0")
    footer_separator.pack(fill="x", pady=(0, 12))

    # Información del footer
    footer_content = tk.Frame(footer_frame, bg="#FFFFFF")
    footer_content.pack(fill="x")

    tk.Label(footer_content, 
             text="Sistema V&C v4.0 • © 2026",
             font=("Inter", 9),
             fg="#4B4B4B", 
             bg="#FFFFFF").pack()

    root.mainloop()